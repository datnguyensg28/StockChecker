# =====================================================
<<<<<<< ours
# STOCKFLOW ONLINE - KIEM TRA KHA NANG XUAT KHO
# Author: DatND5
# Version: 2.0 Streamlit Online
# =====================================================

import io
import datetime
from typing import Dict, Tuple, Any, Optional
=======
# STOCKFLOW CHECKER - KIEM TRA DAM BAO XUAT KHO
# Author: DatND5
# Version: 3.0 Streamlit
# =====================================================

import datetime
import io
from typing import Any, Dict, Optional, Tuple
>>>>>>> theirs

import pandas as pd
import requests
import streamlit as st
<<<<<<< ours
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

=======
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


>>>>>>> theirs
# =====================================================
# PAGE CONFIG
# =====================================================
st.set_page_config(
<<<<<<< ours
    page_title="StockFlow Online",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="expanded",
)

# =====================================================
# CONFIG
# =====================================================
# Cach 1 - Khuyen dung: dat trong Streamlit Secrets:
# MB52_RAW_URL = "https://raw.githubusercontent.com/datnguyensg28/StockChecker/main/data/MB52.XLSX"
#
# Cach 2 - Tam thoi: dien truc tiep URL ben duoi.
DEFAULT_MB52_RAW_URL = "https://raw.githubusercontent.com/datnguyensg28/StockChecker/main/data/MB52.XLSX"

APP_NAME = "StockFlow Online"
APP_SUBTITLE = "Kiểm tra khả năng xuất kho theo tồn kho MB52 mới nhất"
APP_VERSION = "2.0"
=======
    page_title="StockFlow Checker",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="collapsed",
)


# =====================================================
# CONFIG
# =====================================================
DEFAULT_MB52_RAW_URL = "https://raw.githubusercontent.com/datnguyensg28/StockChecker/main/data/MB52.XLSX"
LOCAL_MB52_PATH = "data/MB52.XLSX"

APP_NAME = "StockFlow Checker"
APP_SUBTITLE = "Kiểm tra phiếu xuất kho theo trạng thái thực xuất và tồn kho MB52"
APP_VERSION = "3.0"
>>>>>>> theirs

REQUIRED_MB52_COLUMNS = ["Material", "Plant", "Unrestricted", "WBS Element"]
REQUIRED_ISSUE_COLUMNS = [
    "Request Number",
    "Material Number",
    "Material Description",
    "Plant",
    "Source WBS",
    "Sending Sloc",
    "Functional Location",
    "Transfer Quantity",
]

<<<<<<< ours
REPORT_COLUMNS = [
=======
DETAIL_COLUMNS = [
>>>>>>> theirs
    "Request Number",
    "Material Number",
    "Material Description",
    "Plant",
    "Source WBS",
    "Sending Sloc",
    "Functional Location",
    "Transfer Quantity",
<<<<<<< ours
=======
    "Actual Quantity",
    "Status",
    "Còn thiếu",
    "Tình trạng",
    "Gợi ý xử lý",
]

STOCK_COLUMNS = [
>>>>>>> theirs
    "Tồn kho DA CN",
    "Tồn kho DA Tỉnh",
    "Tồn kho CN",
    "Tồn kho Tỉnh",
    "Tồn kho Khu vực",
<<<<<<< ours
    "Tầng đáp ứng",
    "Report Status",
    "Gợi ý chuyển WBS",
]

=======
]


>>>>>>> theirs
# =====================================================
# CSS
# =====================================================
st.markdown(
    """
    <style>
        .main .block-container {
<<<<<<< ours
            padding-top: 1.5rem;
            padding-bottom: 2rem;
        }
        .hero-card {
            padding: 24px 28px;
            border-radius: 20px;
            background: linear-gradient(135deg, #0f172a 0%, #1e3a8a 55%, #0f766e 100%);
            color: white;
            box-shadow: 0 12px 30px rgba(15, 23, 42, 0.18);
            margin-bottom: 18px;
        }
        .hero-title {
            font-size: 34px;
            font-weight: 800;
            margin-bottom: 4px;
        }
        .hero-subtitle {
            font-size: 16px;
            opacity: 0.92;
        }
        .badge {
            display: inline-block;
            padding: 5px 10px;
            border-radius: 999px;
            background: rgba(255, 255, 255, 0.14);
            border: 1px solid rgba(255, 255, 255, 0.22);
            margin-top: 12px;
            font-size: 13px;
        }
        .section-title {
            margin-top: 10px;
            padding: 12px 16px;
            border-left: 5px solid #2563eb;
            background: #f8fafc;
            border-radius: 12px;
            font-size: 18px;
            font-weight: 700;
            color: #0f172a;
        }
        div[data-testid="stMetric"] {
            background: #ffffff;
            padding: 16px;
            border-radius: 16px;
            border: 1px solid #e5e7eb;
            box-shadow: 0 4px 16px rgba(15, 23, 42, 0.06);
        }
        div[data-testid="stMetricValue"] {
            font-size: 28px;
            font-weight: 800;
        }
        .small-note {
            color: #64748b;
            font-size: 13px;
        }
        .success-pill {
            padding: 3px 8px;
            border-radius: 999px;
            background: #dcfce7;
            color: #166534;
            font-weight: 700;
            font-size: 12px;
        }
        .danger-pill {
            padding: 3px 8px;
            border-radius: 999px;
            background: #fee2e2;
            color: #991b1b;
            font-weight: 700;
            font-size: 12px;
        }
        .stDownloadButton > button {
            border-radius: 12px;
            font-weight: 700;
=======
            padding-top: 1.2rem;
            padding-bottom: 2rem;
            max-width: 1180px;
        }
        .app-header {
            padding: 18px 0 10px 0;
            border-bottom: 1px solid #e5e7eb;
            margin-bottom: 18px;
        }
        .app-title {
            font-size: 32px;
            font-weight: 800;
            color: #111827;
            line-height: 1.15;
        }
        .app-subtitle {
            color: #4b5563;
            margin-top: 6px;
            font-size: 15px;
        }
        .step-title {
            margin: 18px 0 10px 0;
            padding: 12px 14px;
            background: #f9fafb;
            border: 1px solid #e5e7eb;
            border-left: 5px solid #2563eb;
            border-radius: 8px;
            font-weight: 800;
            color: #111827;
        }
        div[data-testid="stMetric"] {
            background: #ffffff;
            padding: 14px 16px;
            border-radius: 8px;
            border: 1px solid #e5e7eb;
        }
        div[data-testid="stMetricValue"] {
            font-size: 26px;
            font-weight: 800;
        }
        .result-card {
            border-radius: 8px;
            padding: 22px 24px;
            margin: 14px 0 16px 0;
            border: 1px solid;
        }
        .result-ok {
            background: #ecfdf5;
            border-color: #86efac;
            color: #14532d;
        }
        .result-bad {
            background: #fff7ed;
            border-color: #fdba74;
            color: #7c2d12;
        }
        .result-headline {
            font-size: 30px;
            font-weight: 900;
            margin-bottom: 8px;
        }
        .result-copy {
            font-size: 17px;
            font-weight: 600;
        }
        .small-note {
            color: #6b7280;
            font-size: 13px;
>>>>>>> theirs
        }
    </style>
    """,
    unsafe_allow_html=True,
)

<<<<<<< ours
# =====================================================
# HELPER FUNCTIONS
# =====================================================
def get_mb52_raw_url() -> str:
    """Lay URL MB52 tu secrets neu co, neu khong dung DEFAULT."""
    try:
        url = st.secrets.get("MB52_RAW_URL", DEFAULT_MB52_RAW_URL)
    except Exception:
        url = DEFAULT_MB52_RAW_URL
    return str(url).strip()
=======

# =====================================================
# HELPERS
# =====================================================
def get_mb52_raw_url() -> str:
    try:
        return str(st.secrets.get("MB52_RAW_URL", DEFAULT_MB52_RAW_URL)).strip()
    except Exception:
        return DEFAULT_MB52_RAW_URL
>>>>>>> theirs


def normalize_key_value(value: Any) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


<<<<<<< ours
def validate_columns(df: pd.DataFrame, required_cols: list, file_label: str) -> None:
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        st.error(f"❌ File {file_label} thiếu cột bắt buộc: {', '.join(missing)}")
        st.stop()


def detect_storage_location_column(df: pd.DataFrame) -> Optional[str]:
    exact_candidates = [c for c in df.columns if str(c).strip().lower() == "storage location"]
=======
def normalize_column_name(value: Any) -> str:
    return str(value).strip().lower()


def stop_with_missing_columns(missing: list[str], file_label: str) -> None:
    st.error(f"❌ File {file_label} thiếu cột bắt buộc: {', '.join(missing)}")
    st.stop()


def validate_columns(df: pd.DataFrame, required_cols: list[str], file_label: str) -> None:
    missing = [col for col in required_cols if col not in df.columns]
    if missing:
        stop_with_missing_columns(missing, file_label)


def detect_storage_location_column(df: pd.DataFrame) -> Optional[str]:
    exact_candidates = [c for c in df.columns if normalize_column_name(c) == "storage location"]
>>>>>>> theirs
    if exact_candidates:
        return exact_candidates[0]

    fuzzy_candidates = [
<<<<<<< ours
        c for c in df.columns
        if "storage" in str(c).lower() and "location" in str(c).lower()
=======
        c
        for c in df.columns
        if "storage" in normalize_column_name(c) and "location" in normalize_column_name(c)
>>>>>>> theirs
    ]
    if fuzzy_candidates:
        return fuzzy_candidates[0]
    return None


<<<<<<< ours
@st.cache_data(ttl=300, show_spinner="🔄 Đang tải MB52 mới nhất từ GitHub...")
def download_mb52_from_github(raw_url: str) -> Tuple[bytes, Dict[str, str]]:
    """Tai file MB52 moi nhat tu GitHub Raw. Cache 5 phut."""
    if not raw_url or "<user>" in raw_url:
        raise ValueError("Chưa cấu hình MB52_RAW_URL. Hãy thay link GitHub Raw thật.")
=======
def detect_column_by_name_or_position(
    df: pd.DataFrame,
    accepted_names: list[str],
    excel_column_index: int,
    display_name: str,
) -> str:
    normalized_names = {name.strip().lower() for name in accepted_names}
    for col in df.columns:
        if normalize_column_name(col) in normalized_names:
            return col

    zero_based_index = excel_column_index - 1
    if len(df.columns) > zero_based_index:
        return df.columns[zero_based_index]

    st.error(
        f"❌ Không tìm thấy cột {display_name}. "
        f"Hãy đặt tên cột là {accepted_names[0]} hoặc đặt đúng vị trí cột Excel."
    )
    st.stop()


def normalize_status(value: Any) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text


def is_exported_status(value: Any) -> bool:
    return normalize_status(value) == "12"


@st.cache_data(ttl=300, show_spinner="Đang tải MB52 mới nhất từ GitHub...")
def download_mb52_from_github(raw_url: str) -> Tuple[bytes, Dict[str, str]]:
    if not raw_url:
        raise ValueError("Chưa cấu hình GitHub Raw URL MB52.")
>>>>>>> theirs

    headers = {
        "Cache-Control": "no-cache",
        "Pragma": "no-cache",
<<<<<<< ours
        "User-Agent": "StockFlow-Online/2.0",
=======
        "User-Agent": "StockFlow-Checker/3.0",
>>>>>>> theirs
    }
    response = requests.get(raw_url, headers=headers, timeout=60)
    response.raise_for_status()

    meta = {
<<<<<<< ours
        "url": raw_url,
        "downloaded_at": datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "content_length": response.headers.get("Content-Length", ""),
        "etag": response.headers.get("ETag", ""),
        "last_modified": response.headers.get("Last-Modified", ""),
=======
        "source": "GitHub - MB52 mới nhất",
        "url": raw_url,
        "loaded_at": datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "last_modified": response.headers.get("Last-Modified", ""),
        "etag": response.headers.get("ETag", ""),
>>>>>>> theirs
    }
    return response.content, meta


<<<<<<< ours
@st.cache_data(show_spinner="🔄 Đang đọc MB52...")
=======
@st.cache_data(show_spinner="Đang đọc MB52 local...")
def read_local_mb52(path: str) -> Tuple[bytes, Dict[str, str]]:
    with open(path, "rb") as file:
        content = file.read()
    meta = {
        "source": "Local - data/MB52.XLSX",
        "url": path,
        "loaded_at": datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "last_modified": "",
        "etag": "",
    }
    return content, meta


@st.cache_data(show_spinner="Đang đọc MB52...")
>>>>>>> theirs
def load_mb52(file_bytes: bytes) -> pd.DataFrame:
    df = pd.read_excel(io.BytesIO(file_bytes))

    sloc_col = detect_storage_location_column(df)
    if not sloc_col:
        st.error("❌ Không tìm thấy cột Storage Location trong MB52.")
        st.stop()
<<<<<<< ours

    if sloc_col != "Storage Location":
        df.rename(columns={sloc_col: "Storage Location"}, inplace=True)
=======
    if sloc_col != "Storage Location":
        df = df.rename(columns={sloc_col: "Storage Location"})
>>>>>>> theirs

    validate_columns(df, REQUIRED_MB52_COLUMNS + ["Storage Location"], "MB52")

    df["Unrestricted"] = pd.to_numeric(df["Unrestricted"], errors="coerce").fillna(0)
<<<<<<< ours

    key_cols = ["Material", "Plant", "Storage Location", "WBS Element"]
    for col in key_cols:
=======
    for col in ["Material", "Plant", "Storage Location", "WBS Element"]:
>>>>>>> theirs
        df[col] = df[col].apply(normalize_key_value)

    return df


<<<<<<< ours
@st.cache_data(show_spinner="🔄 Đang đọc file phiếu xuất...")
def load_issue(file_bytes: bytes) -> pd.DataFrame:
    df = pd.read_excel(io.BytesIO(file_bytes))
    validate_columns(df, REQUIRED_ISSUE_COLUMNS, "phiếu xuất")

    df["Transfer Quantity"] = pd.to_numeric(df["Transfer Quantity"], errors="coerce").fillna(0)
    if "Actual Quantity" in df.columns:
        df["Actual Quantity"] = pd.to_numeric(df["Actual Quantity"], errors="coerce").fillna(0)
    else:
        df["Actual Quantity"] = 0

    key_cols = ["Material Number", "Plant", "Source WBS", "Sending Sloc", "Functional Location"]
    for col in key_cols:
=======
@st.cache_data(show_spinner="Đang đọc file phiếu xuất kho...")
def load_issue(file_bytes: bytes) -> pd.DataFrame:
    df = pd.read_excel(io.BytesIO(file_bytes))
    validate_columns(df, REQUIRED_ISSUE_COLUMNS, "phiếu xuất kho")

    actual_col = detect_column_by_name_or_position(
        df,
        ["Actual Quantity", "Thực xuất"],
        28,  # AB
        "Actual Quantity / Thực xuất",
    )
    status_col = detect_column_by_name_or_position(
        df,
        ["Status"],
        29,  # AC
        "Status",
    )

    if actual_col != "Actual Quantity":
        df = df.rename(columns={actual_col: "Actual Quantity"})
    if status_col != "Status":
        df = df.rename(columns={status_col: "Status"})

    df["Transfer Quantity"] = pd.to_numeric(df["Transfer Quantity"], errors="coerce").fillna(0)
    df["Actual Quantity"] = pd.to_numeric(df["Actual Quantity"], errors="coerce").fillna(0)
    df["Status"] = df["Status"].apply(normalize_status)

    for col in ["Material Number", "Plant", "Source WBS", "Sending Sloc", "Functional Location"]:
>>>>>>> theirs
        df[col] = df[col].apply(normalize_key_value)

    return df


def build_inventory_maps(mb52_raw: pd.DataFrame):
<<<<<<< ours
    map_da_cn = mb52_raw.groupby(
        ["Material", "Plant", "Storage Location", "WBS Element"], as_index=False
    )["Unrestricted"].sum().set_index(
        ["Material", "Plant", "Storage Location", "WBS Element"]
    )["Unrestricted"].to_dict()

    map_da_tinh = mb52_raw.groupby(
        ["Material", "Plant", "WBS Element"], as_index=False
    )["Unrestricted"].sum().set_index(
        ["Material", "Plant", "WBS Element"]
    )["Unrestricted"].to_dict()

    map_cn = mb52_raw.groupby(
        ["Material", "Plant", "Storage Location"], as_index=False
    )["Unrestricted"].sum().set_index(
        ["Material", "Plant", "Storage Location"]
    )["Unrestricted"].to_dict()

    map_tinh = mb52_raw.groupby(
        ["Material", "Plant"], as_index=False
    )["Unrestricted"].sum().set_index(
        ["Material", "Plant"]
    )["Unrestricted"].to_dict()

    map_kv = mb52_raw.groupby(
        ["Material"], as_index=False
    )["Unrestricted"].sum().set_index(
        ["Material"]
    )["Unrestricted"].to_dict()
=======
    map_da_cn = (
        mb52_raw.groupby(["Material", "Plant", "Storage Location", "WBS Element"], as_index=False)["Unrestricted"]
        .sum()
        .set_index(["Material", "Plant", "Storage Location", "WBS Element"])["Unrestricted"]
        .to_dict()
    )
    map_da_tinh = (
        mb52_raw.groupby(["Material", "Plant", "WBS Element"], as_index=False)["Unrestricted"]
        .sum()
        .set_index(["Material", "Plant", "WBS Element"])["Unrestricted"]
        .to_dict()
    )
    map_cn = (
        mb52_raw.groupby(["Material", "Plant", "Storage Location"], as_index=False)["Unrestricted"]
        .sum()
        .set_index(["Material", "Plant", "Storage Location"])["Unrestricted"]
        .to_dict()
    )
    map_tinh = (
        mb52_raw.groupby(["Material", "Plant"], as_index=False)["Unrestricted"]
        .sum()
        .set_index(["Material", "Plant"])["Unrestricted"]
        .to_dict()
    )
    map_kv = mb52_raw.groupby(["Material"], as_index=False)["Unrestricted"].sum().set_index(["Material"])["Unrestricted"].to_dict()
>>>>>>> theirs

    return map_da_cn, map_da_tinh, map_cn, map_tinh, map_kv


def build_sequential_5_layer(issue_df: pd.DataFrame, mb52_raw: pd.DataFrame) -> pd.DataFrame:
    map_da_cn, map_da_tinh, map_cn, map_tinh, map_kv = build_inventory_maps(mb52_raw)

    r = issue_df.copy()
<<<<<<< ours

=======
>>>>>>> theirs
    remain_da_cn = map_da_cn.copy()
    remain_da_tinh = map_da_tinh.copy()
    remain_cn = map_cn.copy()
    remain_tinh = map_tinh.copy()

    r["Tầng đáp ứng"] = ""
    r["Gợi ý chuyển WBS"] = ""
    r["Report Status"] = ""
    r["Thiếu kho"] = False

<<<<<<< ours
    stock_cols = ["Tồn kho DA CN", "Tồn kho DA Tỉnh", "Tồn kho CN", "Tồn kho Tỉnh", "Tồn kho Khu vực"]
    for col in stock_cols:
=======
    for col in STOCK_COLUMNS:
>>>>>>> theirs
        r[col] = 0.0

    for idx, row in r.iterrows():
        qty = float(row["Transfer Quantity"])
        mat = normalize_key_value(row["Material Number"])
        plant = normalize_key_value(row["Plant"])
        sloc = normalize_key_value(row["Sending Sloc"])
        wbs = normalize_key_value(row["Source WBS"])

        da_cn_key = (mat, plant, sloc, wbs)
        da_tinh_key = (mat, plant, wbs)
        cn_key = (mat, plant, sloc)
        tinh_key = (mat, plant)
        kv_key = mat

        r.at[idx, "Tồn kho DA CN"] = remain_da_cn.get(da_cn_key, 0)
        r.at[idx, "Tồn kho DA Tỉnh"] = remain_da_tinh.get(da_tinh_key, 0)
        r.at[idx, "Tồn kho CN"] = remain_cn.get(cn_key, 0)
        r.at[idx, "Tồn kho Tỉnh"] = remain_tinh.get(tinh_key, 0)
        r.at[idx, "Tồn kho Khu vực"] = map_kv.get(kv_key, 0)

        da_cn_qty = remain_da_cn.get(da_cn_key, 0)
<<<<<<< ours

=======
>>>>>>> theirs
        if qty <= da_cn_qty:
            remain_da_cn[da_cn_key] = da_cn_qty - qty
            r.at[idx, "Tầng đáp ứng"] = "Kho DA CN"
            r.at[idx, "Report Status"] = "ĐẢM BẢO"
            continue

        r.at[idx, "Report Status"] = "KHÔNG ĐẢM BẢO"
        r.at[idx, "Thiếu kho"] = True

        if qty <= remain_da_tinh.get(da_tinh_key, 0):
            remain_da_tinh[da_tinh_key] -= qty
<<<<<<< ours
            r.at[idx, "Gợi ý chuyển WBS"] = "🧠 Có thể chuyển từ Kho DA Tỉnh"
        elif qty <= remain_cn.get(cn_key, 0):
            remain_cn[cn_key] -= qty
            r.at[idx, "Gợi ý chuyển WBS"] = "🧠 Có thể chuyển từ Kho CN"
        elif qty <= remain_tinh.get(tinh_key, 0):
            remain_tinh[tinh_key] -= qty
            r.at[idx, "Gợi ý chuyển WBS"] = "🧠 Có thể chuyển từ Kho Tỉnh"
        elif qty <= map_kv.get(kv_key, 0):
            r.at[idx, "Gợi ý chuyển WBS"] = "🧠 Có thể điều chuyển từ Kho Khu vực"
        else:
            r.at[idx, "Gợi ý chuyển WBS"] = "🚚 Thiếu toàn bộ các tầng kho"
=======
            r.at[idx, "Gợi ý chuyển WBS"] = "Có thể chuyển từ Kho DA Tỉnh"
        elif qty <= remain_cn.get(cn_key, 0):
            remain_cn[cn_key] -= qty
            r.at[idx, "Gợi ý chuyển WBS"] = "Có thể chuyển từ Kho CN"
        elif qty <= remain_tinh.get(tinh_key, 0):
            remain_tinh[tinh_key] -= qty
            r.at[idx, "Gợi ý chuyển WBS"] = "Có thể chuyển từ Kho Tỉnh"
        elif qty <= map_kv.get(kv_key, 0):
            r.at[idx, "Gợi ý chuyển WBS"] = "Có thể điều chuyển từ Kho Khu vực"
        else:
            r.at[idx, "Gợi ý chuyển WBS"] = "Thiếu toàn bộ các tầng kho"
>>>>>>> theirs

    return r


<<<<<<< ours
def apply_sidebar_filters(df: pd.DataFrame) -> pd.DataFrame:
    filtered = df.copy()

    st.sidebar.markdown("---")
    st.sidebar.header("🔍 Bộ lọc nhanh")

    filter_material = st.sidebar.text_input("Mã vật tư")

    fl_search = st.sidebar.text_input("Tìm nhanh Functional Location")
    all_fl = sorted(filtered["Functional Location"].dropna().astype(str).unique())
    if fl_search:
        all_fl = [f for f in all_fl if fl_search.lower() in f.lower()]

    filter_fl = st.sidebar.multiselect("Functional Location", all_fl)
    filter_plant = st.sidebar.multiselect("Plant", sorted(filtered["Plant"].dropna().astype(str).unique()))
    filter_status = st.sidebar.multiselect("Tình trạng", ["ĐẢM BẢO", "KHÔNG ĐẢM BẢO"])
    filter_layer = st.sidebar.multiselect("Tầng đáp ứng", sorted(filtered["Tầng đáp ứng"].dropna().astype(str).unique()))

    if filter_material:
        filtered = filtered[filtered["Material Number"].astype(str).str.contains(filter_material, case=False, na=False)]
    if filter_fl:
        filtered = filtered[filtered["Functional Location"].astype(str).isin(filter_fl)]
    if filter_plant:
        filtered = filtered[filtered["Plant"].astype(str).isin(filter_plant)]
    if filter_status:
        filtered = filtered[filtered["Report Status"].isin(filter_status)]
    if filter_layer:
        filtered = filtered[filtered["Tầng đáp ứng"].isin(filter_layer)]

    return filtered


def show_dashboard(report_df: pd.DataFrame) -> None:
    total = len(report_df)
    ok = int((report_df["Report Status"] == "ĐẢM BẢO").sum())
    not_ok = int((report_df["Report Status"] == "KHÔNG ĐẢM BẢO").sum())
    ok_rate = (ok / total * 100) if total else 0
    fl_missing = report_df.loc[report_df["Thiếu kho"], "Functional Location"].nunique()

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Tổng dòng", f"{total:,}")
    c2.metric("Đảm bảo", f"{ok:,}")
    c3.metric("Không đảm bảo", f"{not_ok:,}")
    c4.metric("Tỷ lệ đảm bảo", f"{ok_rate:.1f}%")
    c5.metric("FL thiếu kho", f"{fl_missing:,}")

    st.markdown('<div class="section-title">📈 Tổng quan tình trạng</div>', unsafe_allow_html=True)
    chart_col1, chart_col2 = st.columns(2)

    with chart_col1:
        status_summary = report_df.groupby("Report Status").size().reset_index(name="Số dòng")
        st.bar_chart(status_summary, x="Report Status", y="Số dòng", use_container_width=True)

    with chart_col2:
        plant_summary = report_df[report_df["Thiếu kho"]].groupby("Plant").size().reset_index(name="Số dòng thiếu")
        if not plant_summary.empty:
            st.bar_chart(plant_summary, x="Plant", y="Số dòng thiếu", use_container_width=True)
        else:
            st.success("Không có dữ liệu thiếu kho theo Plant.")


def build_summaries(report_df: pd.DataFrame):
    missing_df = report_df[report_df["Thiếu kho"]].copy()

    summary_fl = missing_df.groupby("Functional Location").size().reset_index(name="Số dòng thiếu kho")
    summary_fl = summary_fl.sort_values("Số dòng thiếu kho", ascending=False)

    summary_material = missing_df.groupby(["Material Number", "Material Description"]).agg(
        **{"Số dòng thiếu kho": ("Material Number", "size"), "Tổng SL yêu cầu": ("Transfer Quantity", "sum")}
    ).reset_index().sort_values("Số dòng thiếu kho", ascending=False)

    summary_plant = missing_df.groupby("Plant").agg(
        **{"Số dòng thiếu kho": ("Plant", "size"), "Tổng SL yêu cầu": ("Transfer Quantity", "sum")}
    ).reset_index().sort_values("Số dòng thiếu kho", ascending=False)

    suggestion = missing_df[[
        "Request Number",
        "Material Number",
        "Material Description",
        "Plant",
        "Source WBS",
        "Sending Sloc",
        "Functional Location",
        "Transfer Quantity",
        "Gợi ý chuyển WBS",
    ]].copy()

    return summary_fl, summary_material, summary_plant, suggestion


def auto_width_worksheet(ws):
    for col_idx, column_cells in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column_cells:
            try:
                cell_length = len(str(cell.value)) if cell.value is not None else 0
                max_length = max(max_length, cell_length)
            except Exception:
                pass
        adjusted_width = min(max(max_length + 2, 12), 45)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width


def format_report_workbook(writer, sheet_names):
    wb = writer.book
    header_fill = PatternFill("solid", fgColor="1E3A8A")
    header_font = Font(color="FFFFFF", bold=True)
    ok_fill = PatternFill("solid", fgColor="DCFCE7")
    bad_fill = PatternFill("solid", fgColor="FEE2E2")
    suggest_fill = PatternFill("solid", fgColor="FEF9C3")
=======
def build_business_conclusion(report_df: pd.DataFrame) -> pd.DataFrame:
    r = report_df.copy()

    exported = r["Status"].apply(is_exported_status)
    enough_actual = r["Actual Quantity"] >= r["Transfer Quantity"]
    enough_mb52 = ~r["Thiếu kho"]

    shortage_by_actual = (r["Transfer Quantity"] - r["Actual Quantity"]).clip(lower=0)
    shortage_by_mb52 = r["Transfer Quantity"].where(~enough_mb52, 0)
    r["Còn thiếu"] = shortage_by_actual.where(shortage_by_actual > 0, shortage_by_mb52).fillna(0)

    r["Tình trạng"] = "Đảm bảo xuất kho"
    r["Gợi ý xử lý"] = "Không cần xử lý thêm"

    not_exported_mask = ~exported
    short_actual_mask = exported & ~enough_actual
    mb52_missing_mask = exported & enough_actual & ~enough_mb52

    r.loc[not_exported_mask, "Tình trạng"] = "Chưa xuất kho"
    r.loc[not_exported_mask, "Gợi ý xử lý"] = "Kiểm tra trạng thái phiếu, thực hiện xuất kho để Status = 12"

    r.loc[short_actual_mask, "Tình trạng"] = "Xuất thiếu"
    r.loc[short_actual_mask, "Gợi ý xử lý"] = "Kiểm tra số lượng thực xuất và xuất bổ sung phần còn thiếu"

    r.loc[mb52_missing_mask, "Tình trạng"] = "Thiếu tồn kho MB52"
    r.loc[mb52_missing_mask, "Gợi ý xử lý"] = r.loc[mb52_missing_mask, "Gợi ý chuyển WBS"].fillna("")
    r.loc[mb52_missing_mask & (r["Gợi ý xử lý"] == ""), "Gợi ý xử lý"] = "Kiểm tra bổ sung tồn kho MB52 hoặc điều chuyển vật tư"

    r["Đảm bảo 100%"] = exported & enough_actual & enough_mb52
    return r


def build_conclusion_sheet(total: int, ok: int, not_ok: int, mb52_meta: Dict[str, str]) -> pd.DataFrame:
    ok_rate = (ok / total * 100) if total else 0
    conclusion = (
        "ĐẢM BẢO XUẤT KHO 100%"
        if total > 0 and not_ok == 0
        else "CHƯA ĐẢM BẢO XUẤT KHO 100%"
    )
    return pd.DataFrame(
        [
            {"Thông tin": "Kết luận", "Giá trị": conclusion},
            {"Thông tin": "Tổng dòng", "Giá trị": total},
            {"Thông tin": "Đã xuất đủ", "Giá trị": ok},
            {"Thông tin": "Chưa đảm bảo", "Giá trị": not_ok},
            {"Thông tin": "Tỷ lệ đảm bảo", "Giá trị": f"{ok_rate:.1f}%"},
            {"Thông tin": "Nguồn MB52", "Giá trị": mb52_meta.get("source", "")},
            {"Thông tin": "MB52 URL/Path", "Giá trị": mb52_meta.get("url", "")},
            {"Thông tin": "Thời điểm kiểm tra", "Giá trị": datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")},
        ]
    )


def auto_width_worksheet(ws) -> None:
    for col_idx, column_cells in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column_cells:
            cell_length = len(str(cell.value)) if cell.value is not None else 0
            max_length = max(max_length, cell_length)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_length + 2, 12), 48)


def format_workbook(writer, sheet_names: list[str]) -> None:
    wb = writer.book
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    bad_fill = PatternFill("solid", fgColor="FFEDD5")
>>>>>>> theirs

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
<<<<<<< ours

=======
>>>>>>> theirs
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

<<<<<<< ours
        if sheet_name == "BaoCaoChiTiet":
            headers = [cell.value for cell in ws[1]]
            status_col = headers.index("Report Status") + 1 if "Report Status" in headers else None
            suggest_col = headers.index("Gợi ý chuyển WBS") + 1 if "Gợi ý chuyển WBS" in headers else None

            for row in range(2, ws.max_row + 1):
                if status_col:
                    status = ws.cell(row=row, column=status_col).value
                    fill = ok_fill if status == "ĐẢM BẢO" else bad_fill if status == "KHÔNG ĐẢM BẢO" else None
                    if fill:
                        for col in range(1, ws.max_column + 1):
                            ws.cell(row=row, column=col).fill = fill
                if suggest_col and ws.cell(row=row, column=suggest_col).value:
                    ws.cell(row=row, column=suggest_col).fill = suggest_fill

        auto_width_worksheet(ws)


def export_excel(report_df: pd.DataFrame, mb52_meta: Dict[str, str]) -> bytes:
    summary_fl, summary_material, summary_plant, suggestion = build_summaries(report_df)
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        info_df = pd.DataFrame([
            {"Thông tin": "Tên phần mềm", "Giá trị": APP_NAME},
            {"Thông tin": "Phiên bản", "Giá trị": APP_VERSION},
            {"Thông tin": "Thời điểm xuất báo cáo", "Giá trị": datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")},
            {"Thông tin": "MB52 GitHub URL", "Giá trị": mb52_meta.get("url", "")},
            {"Thông tin": "MB52 tải lúc", "Giá trị": mb52_meta.get("downloaded_at", "")},
            {"Thông tin": "GitHub Last-Modified", "Giá trị": mb52_meta.get("last_modified", "")},
        ])
        info_df.to_excel(writer, index=False, sheet_name="ThongTin")
        report_df[REPORT_COLUMNS].to_excel(writer, index=False, sheet_name="BaoCaoChiTiet")
        summary_fl.to_excel(writer, index=False, sheet_name="TongHopThieuKho_FL")
        summary_material.to_excel(writer, index=False, sheet_name="TongHopThieuKho_VatTu")
        summary_plant.to_excel(writer, index=False, sheet_name="TongHopTheoPlant")
        suggestion.to_excel(writer, index=False, sheet_name="GoiYDieuChuyen")

        format_report_workbook(
            writer,
            ["ThongTin", "BaoCaoChiTiet", "TongHopThieuKho_FL", "TongHopThieuKho_VatTu", "TongHopTheoPlant", "GoiYDieuChuyen"],
        )
=======
        if sheet_name == "ChiTietChuaDamBao":
            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = bad_fill
        auto_width_worksheet(ws)


def export_excel(full_df: pd.DataFrame, issue_df: pd.DataFrame, mb52_meta: Dict[str, str]) -> bytes:
    total = len(full_df)
    ok = int(full_df["Đảm bảo 100%"].sum())
    not_ok = total - ok
    error_df = full_df.loc[~full_df["Đảm bảo 100%"], DETAIL_COLUMNS].copy()
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        sheet_names = ["KetLuan"]
        build_conclusion_sheet(total, ok, not_ok, mb52_meta).to_excel(writer, index=False, sheet_name="KetLuan")

        if not_ok > 0:
            error_df.to_excel(writer, index=False, sheet_name="ChiTietChuaDamBao")
            error_df[
                [
                    "Request Number",
                    "Material Number",
                    "Material Description",
                    "Plant",
                    "Functional Location",
                    "Còn thiếu",
                    "Tình trạng",
                    "Gợi ý xử lý",
                ]
            ].to_excel(writer, index=False, sheet_name="GoiYXuLy")
            sheet_names.extend(["ChiTietChuaDamBao", "GoiYXuLy"])

        format_workbook(writer, sheet_names)
>>>>>>> theirs

    return output.getvalue()


<<<<<<< ours
# =====================================================
# UI - HEADER
# =====================================================
st.markdown(
    f"""
    <div class="hero-card">
        <div class="hero-title">📦 {APP_NAME}</div>
        <div class="hero-subtitle">{APP_SUBTITLE}</div>
        <div class="badge">Version {APP_VERSION} · MB52 từ GitHub · </div>
=======
def render_result_card(is_all_ok: bool) -> None:
    if is_all_ok:
        st.markdown(
            """
            <div class="result-card result-ok">
                <div class="result-headline">✅ ĐẢM BẢO XUẤT KHO 100%</div>
                <div class="result-copy">Loại phiếu này đã xuất đủ toàn bộ vật tư. Không cần xử lý thêm.</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            """
            <div class="result-card result-bad">
                <div class="result-headline">⚠️ CHƯA ĐẢM BẢO XUẤT KHO 100%</div>
                <div class="result-copy">Chỉ các dòng lỗi hoặc chưa đủ được hiển thị bên dưới.</div>
            </div>
            """,
            unsafe_allow_html=True,
        )


# =====================================================
# UI
# =====================================================
st.markdown(
    f"""
    <div class="app-header">
        <div class="app-title">📦 {APP_NAME}</div>
        <div class="app-subtitle">{APP_SUBTITLE} · Version {APP_VERSION}</div>
>>>>>>> theirs
    </div>
    """,
    unsafe_allow_html=True,
)

<<<<<<< ours
# =====================================================
# SIDEBAR - MB52 SOURCE
# =====================================================
st.sidebar.title("⚙️ Cấu hình")
st.sidebar.caption("Nguồn MB52 mới nhất")

mb52_url = get_mb52_raw_url()
custom_url = st.sidebar.text_input("GitHub Raw URL MB52", value=mb52_url)

if st.sidebar.button("🔄 Tải lại MB52 mới nhất"):
    st.cache_data.clear()
    st.rerun()

try:
    mb52_bytes, mb52_meta = download_mb52_from_github(custom_url)
    mb52_raw = load_mb52(mb52_bytes)
    st.sidebar.success("✅ Đã tải MB52 từ GitHub")
    st.sidebar.caption(f"Tải lúc: {mb52_meta.get('downloaded_at')}")
    if mb52_meta.get("last_modified"):
        st.sidebar.caption(f"GitHub Last-Modified: {mb52_meta.get('last_modified')}")
except Exception as exc:
    st.error(f"❌ Không tải được MB52 từ GitHub: {exc}")
    st.info("Hãy kiểm tra lại GitHub Raw URL hoặc cấu hình biến MB52_RAW_URL trong Streamlit Secrets.")
    st.stop()

# =====================================================
# MAIN - MB52 STATUS
# =====================================================
info1, info2, info3, info4 = st.columns(4)
info1.metric("Dòng MB52", f"{len(mb52_raw):,}")
info2.metric("Mã vật tư", f"{mb52_raw['Material'].nunique():,}")
info3.metric("Plant", f"{mb52_raw['Plant'].nunique():,}")
info4.metric("Storage Location", f"{mb52_raw['Storage Location'].nunique():,}")

st.info(
    f"ℹ️ Đang dùng MB52 mới nhất từ GitHub. Thời điểm app tải file: **{mb52_meta.get('downloaded_at')}**. "
    f"Cache tự làm mới sau 5 phút hoặc bấm **Tải lại MB52 mới nhất** ở sidebar."
)

# =====================================================
# UPLOAD ISSUE FILE
# =====================================================
st.markdown('<div class="section-title">📂 Upload file phiếu xuất kho</div>', unsafe_allow_html=True)
issue_file = st.file_uploader(
    "Chọn file phiếu xuất kho cần kiểm tra",
    type=["xlsx", "xls"],
    help="File cần có các cột: Request Number, Material Number, Plant, Source WBS, Sending Sloc, Functional Location, Transfer Quantity...",
)

if not issue_file:
    st.warning("Vui lòng upload file phiếu xuất kho để bắt đầu kiểm tra.")
    st.stop()

issue_bytes = issue_file.getvalue()
issue_df = load_issue(issue_bytes)

with st.spinner("🧠 Đang kiểm tra tồn kho theo 5 tầng..."):
    full_report = build_sequential_5_layer(issue_df, mb52_raw)

filtered_report = apply_sidebar_filters(full_report)

# =====================================================
# DASHBOARD
# =====================================================
show_dashboard(filtered_report)

# =====================================================
# REPORT TABLE
# =====================================================
st.markdown('<div class="section-title">📊 Báo cáo kiểm tra chi tiết</div>', unsafe_allow_html=True)

view_option_col1, view_option_col2 = st.columns([1, 3])
with view_option_col1:
    only_missing = st.toggle("Chỉ xem dòng thiếu kho", value=False)
with view_option_col2:
    st.caption("Bảng có thể sort trực tiếp trên giao diện. Dùng bộ lọc bên trái để lọc nhanh.")

view_df = filtered_report[filtered_report["Thiếu kho"]].copy() if only_missing else filtered_report.copy()

st.dataframe(
    view_df[REPORT_COLUMNS],
    use_container_width=True,
    height=520,
    hide_index=True,
    column_config={
        "Report Status": st.column_config.TextColumn("Report Status", width="medium"),
        "Gợi ý chuyển WBS": st.column_config.TextColumn("Gợi ý chuyển WBS", width="large"),
        "Transfer Quantity": st.column_config.NumberColumn("Transfer Quantity", format="%.2f"),
    },
)

# =====================================================
# SUMMARY TABLES
# =====================================================
summary_fl, summary_material, summary_plant, suggestion_df = build_summaries(filtered_report)

st.markdown('<div class="section-title">📌 Tổng hợp thiếu kho</div>', unsafe_allow_html=True)
tab1, tab2, tab3, tab4 = st.tabs(["Theo FL", "Theo vật tư", "Theo Plant", "Gợi ý điều chuyển"])

with tab1:
    st.dataframe(summary_fl, use_container_width=True, hide_index=True, height=300)
with tab2:
    st.dataframe(summary_material, use_container_width=True, hide_index=True, height=300)
with tab3:
    st.dataframe(summary_plant, use_container_width=True, hide_index=True, height=300)
with tab4:
    st.dataframe(suggestion_df, use_container_width=True, hide_index=True, height=300)

# =====================================================
# EXPORT
# =====================================================
st.markdown('<div class="section-title">📤 Export báo cáo</div>', unsafe_allow_html=True)
export_bytes = export_excel(filtered_report, mb52_meta)
file_time = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

st.download_button(
    label="⬇️ Tải báo cáo Excel chuyên nghiệp",
    data=export_bytes,
    file_name=f"StockFlow_BaoCao_KiemTra_TonKho_{file_time}.xlsx",
=======
st.markdown('<div class="step-title">Bước 1: Chọn nguồn MB52</div>', unsafe_allow_html=True)
source_options = [
    "GitHub - MB52 mới nhất",
    "Local - data/MB52.XLSX",
    "Upload MB52 tạm thời",
]
mb52_source = st.radio("Nguồn dữ liệu MB52", source_options, horizontal=True, label_visibility="collapsed")

mb52_bytes: Optional[bytes] = None
mb52_meta: Dict[str, str] = {}

col_source, col_refresh = st.columns([4, 1])
with col_source:
    if mb52_source == "GitHub - MB52 mới nhất":
        raw_url = st.text_input("GitHub Raw URL MB52", value=get_mb52_raw_url())
        try:
            mb52_bytes, mb52_meta = download_mb52_from_github(raw_url)
        except Exception as exc:
            st.error(f"❌ Không tải được MB52 từ GitHub: {exc}")
            st.stop()
    elif mb52_source == "Local - data/MB52.XLSX":
        try:
            mb52_bytes, mb52_meta = read_local_mb52(LOCAL_MB52_PATH)
        except Exception as exc:
            st.error(f"❌ Không đọc được file local {LOCAL_MB52_PATH}: {exc}")
            st.stop()
    else:
        upload_mb52 = st.file_uploader("Upload MB52 tạm thời", type=["xlsx", "xls"], key="mb52_upload")
        if not upload_mb52:
            st.info("Vui lòng upload file MB52 để tiếp tục.")
            st.stop()
        mb52_bytes = upload_mb52.getvalue()
        mb52_meta = {
            "source": "Upload MB52 tạm thời",
            "url": upload_mb52.name,
            "loaded_at": datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
            "last_modified": "",
            "etag": "",
        }

with col_refresh:
    st.write("")
    st.write("")
    if st.button("🔄 Làm mới MB52", use_container_width=True):
        st.cache_data.clear()
        st.rerun()

mb52_raw = load_mb52(mb52_bytes)
st.success(
    f"Đã sẵn sàng MB52: {len(mb52_raw):,} dòng · "
    f"{mb52_raw['Material'].nunique():,} mã vật tư · "
    f"nguồn {mb52_meta.get('source', '')}"
)

st.markdown('<div class="step-title">Bước 2: Upload phiếu xuất kho</div>', unsafe_allow_html=True)
issue_file = st.file_uploader(
    "Chọn file phiếu xuất kho",
    type=["xlsx", "xls"],
    help="File cần có Transfer Quantity, Actual Quantity ở cột AB hoặc theo tên cột, và Status ở cột AC hoặc theo tên cột.",
)
if not issue_file:
    st.info("Upload phiếu xuất kho để phần mềm kết luận ngay.")
    st.stop()

issue_df = load_issue(issue_file.getvalue())

with st.spinner("Đang kiểm tra trạng thái thực xuất và tồn kho MB52 theo 5 tầng..."):
    stock_report = build_sequential_5_layer(issue_df, mb52_raw)
    final_report = build_business_conclusion(stock_report)

total_lines = len(final_report)
ok_lines = int(final_report["Đảm bảo 100%"].sum())
not_ok_lines = total_lines - ok_lines
ok_rate = (ok_lines / total_lines * 100) if total_lines else 0
is_all_ok = total_lines > 0 and not_ok_lines == 0

st.markdown('<div class="step-title">Bước 3: Xem kết luận</div>', unsafe_allow_html=True)
metric1, metric2, metric3, metric4 = st.columns(4)
metric1.metric("Tổng dòng", f"{total_lines:,}")
metric2.metric("Đã xuất đủ", f"{ok_lines:,}")
metric3.metric("Chưa đảm bảo", f"{not_ok_lines:,}")
metric4.metric("Tỷ lệ đảm bảo", f"{ok_rate:.1f}%")

render_result_card(is_all_ok)

error_df = final_report.loc[~final_report["Đảm bảo 100%"], DETAIL_COLUMNS].copy()

if not is_all_ok:
    error_counts = error_df["Tình trạng"].value_counts().rename_axis("Tình trạng").reset_index(name="Số dòng")
    st.dataframe(error_counts, use_container_width=True, hide_index=True, height=150)

    st.dataframe(
        error_df,
        use_container_width=True,
        hide_index=True,
        height=430,
        column_config={
            "Transfer Quantity": st.column_config.NumberColumn("Transfer Quantity", format="%.2f"),
            "Actual Quantity": st.column_config.NumberColumn("Actual Quantity", format="%.2f"),
            "Còn thiếu": st.column_config.NumberColumn("Còn thiếu", format="%.2f"),
            "Gợi ý xử lý": st.column_config.TextColumn("Gợi ý xử lý", width="large"),
        },
    )

export_bytes = export_excel(final_report, issue_df, mb52_meta)
file_time = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
st.download_button(
    label="⬇️ Tải kết quả Excel",
    data=export_bytes,
    file_name=f"StockFlow_KetQua_XuatKho_{file_time}.xlsx",
>>>>>>> theirs
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
)

<<<<<<< ours
st.caption("StockFlow Online · Một sản phẩm hỗ trợ kiểm tra tồn kho miễn phí từ DatND5")
=======
st.caption("StockFlow Checker · Người dùng upload phiếu, phần mềm trả lời ngay: đảm bảo 100% hoặc thiếu dòng nào, vì sao, xử lý thế nào.")
>>>>>>> theirs
