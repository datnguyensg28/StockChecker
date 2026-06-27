# =====================================================
# STOCKFLOW CHECKER - KIEM TRA DAM BAO XUAT KHO
# Author: DatND5
# Version: 3.0 Streamlit
# =====================================================

import datetime
import io
from typing import Any, Dict, Optional, Tuple

import pandas as pd
import requests
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# =====================================================
# PAGE CONFIG
# =====================================================
st.set_page_config(
    page_title="StockFlow Checker",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="collapsed",
)


# =====================================================
# CONFIG
# =====================================================
DEFAULT_MB52_RAW_URL = "https://raw.githubusercontent.com/datnguyensg28/StockChecker/main/data/MB52.XLSX"
LOCAL_MB52_PATH = "data/MB52.XLSX"

APP_NAME = "StockFlow Checker"
APP_SUBTITLE = "Kiểm tra phiếu xuất kho theo trạng thái thực xuất và tồn kho MB52"
APP_VERSION = "3.0"

REQUIRED_MB52_COLUMNS = ["Material", "Plant", "Unrestricted", "WBS Element"]
REQUIRED_ISSUE_COLUMNS = [
    "Request Number",
    "Material Number",
    "Material Description",
    "Plant",
    "Source WBS",
    "Sending Sloc",
    "Functional Location",
    "Transfer Quantity",
]

DETAIL_COLUMNS = [
    "Request Number",
    "Material Number",
    "Material Description",
    "Plant",
    "Source WBS",
    "Sending Sloc",
    "Functional Location",
    "Transfer Quantity",
    "Actual Quantity",
    "Status",
    "Kh\u00f3a ki\u1ec3m tra MB52",
    "S\u1ed1 d\u00f2ng MB52 kh\u1edbp",
    "T\u1ed3n kho \u0111\u00fang kh\u00f3a MB52",
    "Số lượng cần xử lý",
    "Còn thiếu",
    "Tình trạng",
    "Gợi ý xử lý",
]

STOCK_COLUMNS = [
    "Tồn kho DA CN",
    "Tồn kho DA Tỉnh",
    "Tồn kho CN",
    "Tồn kho Tỉnh",
    "Tồn kho Khu vực",
]

STOCK_DETAIL_COLUMNS = DETAIL_COLUMNS + [
    "Tầng đáp ứng",
    "Tồn kho DA CN",
    "Tồn kho DA Tỉnh",
    "Tồn kho CN",
    "Tồn kho Tỉnh",
    "Tồn kho Khu vực",
    "Gợi ý chuyển WBS",
    "Report Status",
]


# =====================================================
# CSS
# =====================================================
st.markdown(
    """
    <style>
        .main .block-container {
            padding-top: 1.2rem;
            padding-bottom: 2rem;
            max-width: 1180px;
        }
        .app-header {
            padding: 18px 0 10px 0;
            border-bottom: 1px solid #e5e7eb;
            margin-bottom: 18px;
        }
        .app-title {
            font-size: 32px;
            font-weight: 800;
            color: #111827;
            line-height: 1.15;
        }
        .app-subtitle {
            color: #4b5563;
            margin-top: 6px;
            font-size: 15px;
        }
        .step-title {
            margin: 18px 0 10px 0;
            padding: 12px 14px;
            background: #f9fafb;
            border: 1px solid #e5e7eb;
            border-left: 5px solid #2563eb;
            border-radius: 8px;
            font-weight: 800;
            color: #111827;
        }
        div[data-testid="stMetric"] {
            background: #ffffff;
            padding: 14px 16px;
            border-radius: 8px;
            border: 1px solid #e5e7eb;
        }
        div[data-testid="stMetricValue"] {
            font-size: 26px;
            font-weight: 800;
        }
        .result-card {
            border-radius: 8px;
            padding: 22px 24px;
            margin: 14px 0 16px 0;
            border: 1px solid;
        }
        .result-ok {
            background: #ecfdf5;
            border-color: #86efac;
            color: #14532d;
        }
        .result-bad {
            background: #fff7ed;
            border-color: #fdba74;
            color: #7c2d12;
        }
        .result-headline {
            font-size: 30px;
            font-weight: 900;
            margin-bottom: 8px;
        }
        .result-copy {
            font-size: 17px;
            font-weight: 600;
        }
        .small-note {
            color: #6b7280;
            font-size: 13px;
        }
    </style>
    """,
    unsafe_allow_html=True,
)


# =====================================================
# HELPERS
# =====================================================
def get_mb52_raw_url() -> str:
    try:
        return str(st.secrets.get("MB52_RAW_URL", DEFAULT_MB52_RAW_URL)).strip()
    except Exception:
        return DEFAULT_MB52_RAW_URL


def normalize_key_value(value: Any, strip_leading_zeros: bool = False) -> str:
    if pd.isna(value):
        return ""

    if isinstance(value, int):
        text = str(value)
    elif isinstance(value, float):
        text = str(int(value)) if value.is_integer() else str(value).strip()
    else:
        text = str(value).replace("\u00a0", " ").strip()
        if text.endswith(".0"):
            numeric_part = text[:-2]
            if numeric_part.replace("-", "", 1).isdigit():
                text = numeric_part

    text = " ".join(text.split()).upper()
    if strip_leading_zeros and text.isdigit():
        text = text.lstrip("0") or "0"
    return text


def normalize_material_key(value: Any) -> str:
    return normalize_key_value(value, strip_leading_zeros=True)


def normalize_sloc_key(value: Any) -> str:
    return normalize_key_value(value, strip_leading_zeros=True)


def normalize_wbs_key(value: Any) -> str:
    return normalize_key_value(value, strip_leading_zeros=False)


def normalize_column_name(value: Any) -> str:
    return str(value).strip().lower()


def stop_with_missing_columns(missing: list[str], file_label: str) -> None:
    st.error(f"❌ File {file_label} thiếu cột bắt buộc: {', '.join(missing)}")
    st.stop()


def validate_columns(df: pd.DataFrame, required_cols: list[str], file_label: str) -> None:
    missing = [col for col in required_cols if col not in df.columns]
    if missing:
        stop_with_missing_columns(missing, file_label)


def detect_storage_location_column(df: pd.DataFrame) -> Optional[str]:
    exact_candidates = [c for c in df.columns if normalize_column_name(c) == "storage location"]
    if exact_candidates:
        return exact_candidates[0]

    fuzzy_candidates = [
        c
        for c in df.columns
        if "storage" in normalize_column_name(c) and "location" in normalize_column_name(c)
    ]
    if fuzzy_candidates:
        return fuzzy_candidates[0]
    return None


def detect_column_by_name_or_position(
    df: pd.DataFrame,
    accepted_names: list[str],
    excel_column_index: int,
    display_name: str,
) -> str:
    normalized_names = {name.strip().lower() for name in accepted_names}
    for col in df.columns:
        if normalize_column_name(col) in normalized_names:
            return col

    zero_based_index = excel_column_index - 1
    if len(df.columns) > zero_based_index:
        return df.columns[zero_based_index]

    st.error(
        f"❌ Không tìm thấy cột {display_name}. "
        f"Hãy đặt tên cột là {accepted_names[0]} hoặc đặt đúng vị trí cột Excel."
    )
    st.stop()


def normalize_status(value: Any) -> str:
    text = normalize_key_value(value, strip_leading_zeros=True)
    return text


def is_exported_status(value: Any) -> bool:
    return normalize_status(value) == "12"


@st.cache_data(ttl=300, show_spinner="Đang tải MB52 mới nhất từ GitHub...")
def download_mb52_from_github(raw_url: str) -> Tuple[bytes, Dict[str, str]]:
    if not raw_url:
        raise ValueError("Chưa cấu hình GitHub Raw URL MB52.")

    headers = {
        "Cache-Control": "no-cache",
        "Pragma": "no-cache",
        "User-Agent": "StockFlow-Checker/3.0",
    }
    response = requests.get(raw_url, headers=headers, timeout=60)
    response.raise_for_status()

    meta = {
        "source": "GitHub - MB52 mới nhất",
        "url": raw_url,
        "loaded_at": datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "last_modified": response.headers.get("Last-Modified", ""),
        "etag": response.headers.get("ETag", ""),
    }
    return response.content, meta


@st.cache_data(show_spinner="Đang đọc MB52 local...")
def read_local_mb52(path: str) -> Tuple[bytes, Dict[str, str]]:
    with open(path, "rb") as file:
        content = file.read()
    meta = {
        "source": "Local - data/MB52.XLSX",
        "url": path,
        "loaded_at": datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "last_modified": "",
        "etag": "",
    }
    return content, meta


@st.cache_data(show_spinner="Đang đọc MB52...")
def load_mb52(file_bytes: bytes) -> pd.DataFrame:
    df = pd.read_excel(io.BytesIO(file_bytes))

    sloc_col = detect_storage_location_column(df)
    if not sloc_col:
        st.error("❌ Không tìm thấy cột Storage Location trong MB52.")
        st.stop()
    if sloc_col != "Storage Location":
        df = df.rename(columns={sloc_col: "Storage Location"})

    validate_columns(df, REQUIRED_MB52_COLUMNS + ["Storage Location"], "MB52")

    df["Unrestricted"] = pd.to_numeric(df["Unrestricted"], errors="coerce").fillna(0)
    df["Material"] = df["Material"].apply(normalize_material_key)
    df["Plant"] = df["Plant"].apply(normalize_key_value)
    df["Storage Location"] = df["Storage Location"].apply(normalize_sloc_key)
    df["WBS Element"] = df["WBS Element"].apply(normalize_wbs_key)

    return df


@st.cache_data(show_spinner="Đang đọc file phiếu xuất kho...")
def load_issue(file_bytes: bytes) -> pd.DataFrame:
    df = pd.read_excel(io.BytesIO(file_bytes))
    validate_columns(df, REQUIRED_ISSUE_COLUMNS, "phiếu xuất kho")

    actual_col = detect_column_by_name_or_position(
        df,
        ["Actual Quantity", "Thực xuất"],
        28,  # AB
        "Actual Quantity / Thực xuất",
    )
    status_col = detect_column_by_name_or_position(
        df,
        ["Status"],
        29,  # AC
        "Status",
    )

    if actual_col != "Actual Quantity":
        df = df.rename(columns={actual_col: "Actual Quantity"})
    if status_col != "Status":
        df = df.rename(columns={status_col: "Status"})

    df["Transfer Quantity"] = pd.to_numeric(df["Transfer Quantity"], errors="coerce").fillna(0)
    df["Actual Quantity"] = pd.to_numeric(df["Actual Quantity"], errors="coerce").fillna(0)
    df["Status"] = df["Status"].apply(normalize_status)

    df["Material Number"] = df["Material Number"].apply(normalize_material_key)
    df["Plant"] = df["Plant"].apply(normalize_key_value)
    df["Source WBS"] = df["Source WBS"].apply(normalize_wbs_key)
    df["Sending Sloc"] = df["Sending Sloc"].apply(normalize_sloc_key)
    df["Functional Location"] = df["Functional Location"].apply(normalize_key_value)

    return df


COL_LAYER = "\u0054\u1ea7ng \u0111\u00e1p \u1ee9ng"
COL_SUGGEST_TRANSFER = "G\u1ee3i \u00fd chuy\u1ec3n WBS"
COL_MISSING_STOCK = "Thi\u1ebfu kho"
COL_PROCESS_QTY = "S\u1ed1 l\u01b0\u1ee3ng c\u1ea7n x\u1eed l\u00fd"
COL_SHORTAGE = "C\u00f2n thi\u1ebfu"
COL_BUSINESS_STATUS = "T\u00ecnh tr\u1ea1ng"
COL_ACTION = "G\u1ee3i \u00fd x\u1eed l\u00fd"
COL_OK = "\u0110\u1ea3m b\u1ea3o 100%"
COL_CHECK_KEY = "Kh\u00f3a ki\u1ec3m tra MB52"
COL_MATCHED_ROWS = "S\u1ed1 d\u00f2ng MB52 kh\u1edbp"
COL_DIRECT_STOCK = "T\u1ed3n kho \u0111\u00fang kh\u00f3a MB52"

STOCK_CHECK_STATUSES = {"1", "5", "9"}
EXPORTED_STATUS = "12"


def join_unique(values: pd.Series, limit: int = 8) -> str:
    normalized = []
    for value in values.dropna().astype(str):
        text = value.strip()
        if text and text not in normalized:
            normalized.append(text)
    if len(normalized) > limit:
        return ", ".join(normalized[:limit]) + f", +{len(normalized) - limit}"
    return ", ".join(normalized)


def stock_sum_by_mask(mb52_raw: pd.DataFrame, mask: pd.Series) -> float:
    return float(mb52_raw.loc[mask, "Unrestricted"].sum())


def source_summary(source_rows: pd.DataFrame, limit: int = 3) -> str:
    if source_rows.empty:
        return ""
    grouped = (
        source_rows.groupby(["Plant", "Storage Location", "WBS Element"], as_index=False)["Unrestricted"]
        .sum()
        .sort_values("Unrestricted", ascending=False)
    )
    parts = []
    for _, source in grouped.head(limit).iterrows():
        parts.append(
            f"Plant {source['Plant']} / Sloc {source['Storage Location']} / WBS {source['WBS Element']} "
            f"({float(source['Unrestricted']):,.2f})"
        )
    return "; ".join(parts)


def calculate_stock_layers(mb52_raw: pd.DataFrame, mat: str, plant: str, sloc: str, wbs: str, qty: float) -> Dict[str, Any]:
    mat_mask = mb52_raw["Material"] == mat
    plant_mask = mb52_raw["Plant"] == plant
    sloc_mask = mb52_raw["Storage Location"] == sloc
    wbs_mask = mb52_raw["WBS Element"] == wbs

    da_cn_mask = mat_mask & plant_mask & sloc_mask & wbs_mask
    da_tinh_mask = mat_mask & plant_mask & wbs_mask
    cn_mask = mat_mask & plant_mask & sloc_mask
    tinh_mask = mat_mask & plant_mask
    kv_mask = mat_mask

    da_cn_qty = stock_sum_by_mask(mb52_raw, da_cn_mask)
    da_tinh_qty = stock_sum_by_mask(mb52_raw, da_tinh_mask)
    cn_qty = stock_sum_by_mask(mb52_raw, cn_mask)
    tinh_qty = stock_sum_by_mask(mb52_raw, tinh_mask)
    kv_qty = stock_sum_by_mask(mb52_raw, kv_mask)

    result = {
        "T\u1ed3n kho DA CN": da_cn_qty,
        "T\u1ed3n kho DA T\u1ec9nh": da_tinh_qty,
        "T\u1ed3n kho CN": cn_qty,
        "T\u1ed3n kho T\u1ec9nh": tinh_qty,
        "T\u1ed3n kho Khu v\u1ef1c": kv_qty,
        COL_MATCHED_ROWS: int(da_cn_mask.sum()),
        COL_DIRECT_STOCK: da_cn_qty,
        COL_LAYER: "Kh\u00f4ng \u0111\u1ee7 5 t\u1ea7ng",
        COL_SUGGEST_TRANSFER: "Thi\u1ebfu to\u00e0n b\u1ed9 c\u00e1c t\u1ea7ng kho",
    }

    if qty <= da_cn_qty:
        result[COL_LAYER] = "Kho DA CN"
        result[COL_SUGGEST_TRANSFER] = "\u0110\u1ee7 t\u1ed3n kho \u0111\u00fang kho chi nh\u00e1nh v\u00e0 \u0111\u00fang WBS"
    elif qty <= da_tinh_qty:
        sources = source_summary(mb52_raw.loc[da_tinh_mask & ~sloc_mask & (mb52_raw["Unrestricted"] > 0)])
        if not sources:
            sources = source_summary(mb52_raw.loc[da_tinh_mask & (mb52_raw["Unrestricted"] > 0)])
        result[COL_LAYER] = "Kho DA T\u1ec9nh"
        result[COL_SUGGEST_TRANSFER] = f"C\u00f3 th\u1ec3 chuy\u1ec3n kho chi nh\u00e1nh trong c\u00f9ng WBS t\u1eeb {sources}"
    elif qty <= cn_qty:
        sources = source_summary(mb52_raw.loc[cn_mask & ~wbs_mask & (mb52_raw["Unrestricted"] > 0)])
        if not sources:
            sources = source_summary(mb52_raw.loc[cn_mask & (mb52_raw["Unrestricted"] > 0)])
        result[COL_LAYER] = "Kho CN"
        result[COL_SUGGEST_TRANSFER] = f"C\u00f3 th\u1ec3 chuy\u1ec3n d\u1ef1 \u00e1n/WBS t\u1ea1i c\u00f9ng kho chi nh\u00e1nh t\u1eeb {sources}"
    elif qty <= tinh_qty:
        sources = source_summary(mb52_raw.loc[tinh_mask & ~(sloc_mask & wbs_mask) & (mb52_raw["Unrestricted"] > 0)])
        if not sources:
            sources = source_summary(mb52_raw.loc[tinh_mask & (mb52_raw["Unrestricted"] > 0)])
        result[COL_LAYER] = "Kho T\u1ec9nh"
        result[COL_SUGGEST_TRANSFER] = f"C\u00f3 th\u1ec3 chuy\u1ec3n kho/chuy\u1ec3n d\u1ef1 \u00e1n trong c\u00f9ng Plant t\u1eeb {sources}"
    elif qty <= kv_qty:
        sources = source_summary(mb52_raw.loc[kv_mask & ~plant_mask & (mb52_raw["Unrestricted"] > 0)])
        if not sources:
            sources = source_summary(mb52_raw.loc[kv_mask & (mb52_raw["Unrestricted"] > 0)])
        result[COL_LAYER] = "Kho Khu v\u1ef1c"
        result[COL_SUGGEST_TRANSFER] = f"C\u00f3 th\u1ec3 \u0111i\u1ec1u chuy\u1ec3n li\u00ean Plant/khu v\u1ef1c t\u1eeb {sources}"

    return result


def build_pending_stock_report(issue_df: pd.DataFrame, mb52_raw: pd.DataFrame) -> pd.DataFrame:
    pending = issue_df[issue_df["Status"].isin(STOCK_CHECK_STATUSES)].copy()
    if pending.empty:
        return pd.DataFrame()

    group_cols = ["Material Number", "Plant", "Sending Sloc", "Source WBS"]
    grouped = (
        pending.groupby(group_cols, as_index=False)
        .agg(
            **{
                "Request Number": ("Request Number", join_unique),
                "Material Description": ("Material Description", "first"),
                "Functional Location": ("Functional Location", join_unique),
                "Transfer Quantity": ("Transfer Quantity", "sum"),
                "Actual Quantity": ("Actual Quantity", "sum"),
                "Status": ("Status", join_unique),
            }
        )
    )

    records = []
    for _, row in grouped.iterrows():
        mat = normalize_material_key(row["Material Number"])
        plant = normalize_key_value(row["Plant"])
        sloc = normalize_sloc_key(row["Sending Sloc"])
        wbs = normalize_wbs_key(row["Source WBS"])
        qty = float(row["Transfer Quantity"])

        layers = calculate_stock_layers(mb52_raw, mat, plant, sloc, wbs, qty)
        direct_stock = float(layers["T\u1ed3n kho DA CN"])
        is_ok = qty <= direct_stock
        shortage = max(qty - direct_stock, 0)

        record = {
            "Request Number": row["Request Number"],
            "Material Number": mat,
            "Material Description": row["Material Description"],
            "Plant": plant,
            "Source WBS": wbs,
            "Sending Sloc": sloc,
            "Functional Location": row["Functional Location"],
            "Transfer Quantity": qty,
            "Actual Quantity": float(row["Actual Quantity"]),
            "Status": row["Status"],
            COL_CHECK_KEY: f"Material={mat} | Plant={plant} | Sloc={sloc} | WBS={wbs}",
            COL_MATCHED_ROWS: layers[COL_MATCHED_ROWS],
            COL_DIRECT_STOCK: direct_stock,
            COL_PROCESS_QTY: qty,
            COL_SHORTAGE: shortage,
            COL_BUSINESS_STATUS: "Status 1/5/9 - \u0111\u1ee7 t\u1ed3n kho" if is_ok else "Status 1/5/9 - kh\u00f4ng \u0111\u1ee7 t\u1ed3n kho",
            COL_ACTION: "\u0110\u1ee7 t\u1ed3n kho MB52 \u0111\u00fang Material/Plant/Sloc/WBS" if is_ok else layers[COL_SUGGEST_TRANSFER],
            COL_LAYER: "Kho DA CN" if is_ok else layers[COL_LAYER],
            COL_SUGGEST_TRANSFER: layers[COL_SUGGEST_TRANSFER],
            "Report Status": "\u0110\u1ea2M B\u1ea2O" if is_ok else "KH\u00d4NG \u0110\u1ea2M B\u1ea2O",
            COL_MISSING_STOCK: not is_ok,
            COL_OK: is_ok,
        }
        for stock_col in STOCK_COLUMNS:
            record[stock_col] = layers[stock_col]
        records.append(record)

    return pd.DataFrame(records)


def build_exported_status_report(issue_df: pd.DataFrame) -> pd.DataFrame:
    exported = issue_df[issue_df["Status"] == EXPORTED_STATUS].copy()
    if exported.empty:
        return pd.DataFrame()

    records = []
    for _, row in exported.iterrows():
        transfer_qty = float(row["Transfer Quantity"])
        actual_qty = float(row["Actual Quantity"])
        is_equal = abs(transfer_qty - actual_qty) < 1e-9
        shortage = max(transfer_qty - actual_qty, 0)

        if is_equal:
            status_text = "Status 12 - \u0111\u00e3 xu\u1ea5t \u0111\u1ee7"
            action = "Kh\u00f4ng c\u1ea7n x\u1eed l\u00fd th\u00eam"
        elif actual_qty < transfer_qty:
            status_text = "Status 12 - xu\u1ea5t thi\u1ebfu"
            action = "Ki\u1ec3m tra Actual Quantity v\u00e0 xu\u1ea5t b\u1ed5 sung ph\u1ea7n c\u00f2n thi\u1ebfu"
        else:
            status_text = "Status 12 - xu\u1ea5t d\u01b0 so v\u1edbi y\u00eau c\u1ea7u"
            action = "Ki\u1ec3m tra l\u1ea1i Actual Quantity v\u00e0 phi\u1ebfu xu\u1ea5t kho"

        record = {
            "Request Number": row["Request Number"],
            "Material Number": normalize_material_key(row["Material Number"]),
            "Material Description": row["Material Description"],
            "Plant": normalize_key_value(row["Plant"]),
            "Source WBS": normalize_wbs_key(row["Source WBS"]),
            "Sending Sloc": normalize_sloc_key(row["Sending Sloc"]),
            "Functional Location": normalize_key_value(row["Functional Location"]),
            "Transfer Quantity": transfer_qty,
            "Actual Quantity": actual_qty,
            "Status": row["Status"],
            COL_CHECK_KEY: "Status = 12, kh\u00f4ng ki\u1ec3m tra MB52",
            COL_MATCHED_ROWS: 0,
            COL_DIRECT_STOCK: 0.0,
            COL_PROCESS_QTY: shortage,
            COL_SHORTAGE: shortage,
            COL_BUSINESS_STATUS: status_text,
            COL_ACTION: action,
            COL_LAYER: "\u0110\u00e3 xu\u1ea5t kho",
            COL_SUGGEST_TRANSFER: "Status = 12, kh\u00f4ng t\u00ednh t\u1ed3n kho/chuy\u1ec3n kho",
            "Report Status": "\u0110\u1ea2M B\u1ea2O" if is_equal else "KH\u00d4NG \u0110\u1ea2M B\u1ea2O",
            COL_MISSING_STOCK: False,
            COL_OK: is_equal,
        }
        for stock_col in STOCK_COLUMNS:
            record[stock_col] = 0.0
        records.append(record)

    return pd.DataFrame(records)


def build_sequential_5_layer(issue_df: pd.DataFrame, mb52_raw: pd.DataFrame) -> pd.DataFrame:
    pending_report = build_pending_stock_report(issue_df, mb52_raw)
    exported_report = build_exported_status_report(issue_df)
    reports = [df for df in [pending_report, exported_report] if not df.empty]
    if not reports:
        return pd.DataFrame(columns=STOCK_DETAIL_COLUMNS + [COL_OK])
    return pd.concat(reports, ignore_index=True, sort=False)


def build_business_conclusion(report_df: pd.DataFrame) -> pd.DataFrame:
    return report_df.copy()


def build_conclusion_sheet(total: int, ok: int, not_ok: int, mb52_meta: Dict[str, str]) -> pd.DataFrame:
    ok_rate = (ok / total * 100) if total else 0
    conclusion = (
        "ĐẢM BẢO XUẤT KHO 100%"
        if total > 0 and not_ok == 0
        else "CHƯA ĐẢM BẢO XUẤT KHO 100%"
    )
    return pd.DataFrame(
        [
            {"Thông tin": "Kết luận", "Giá trị": conclusion},
            {"Thông tin": "Tổng dòng", "Giá trị": total},
            {"Thông tin": "Đã xuất đủ", "Giá trị": ok},
            {"Thông tin": "Chưa đảm bảo", "Giá trị": not_ok},
            {"Thông tin": "Tỷ lệ đảm bảo", "Giá trị": f"{ok_rate:.1f}%"},
            {"Thông tin": "Nguồn MB52", "Giá trị": mb52_meta.get("source", "")},
            {"Thông tin": "MB52 URL/Path", "Giá trị": mb52_meta.get("url", "")},
            {"Thông tin": "Thời điểm kiểm tra", "Giá trị": datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")},
        ]
    )


def build_stock_summaries(report_df: pd.DataFrame):
    missing_stock_df = report_df[report_df["Thiếu kho"]].copy()

    if missing_stock_df.empty:
        empty_fl = pd.DataFrame(columns=["Functional Location", "Số dòng thiếu kho"])
        empty_material = pd.DataFrame(columns=["Material Number", "Material Description", "Số dòng thiếu kho", "Tổng SL yêu cầu"])
        empty_plant = pd.DataFrame(columns=["Plant", "Số dòng thiếu kho", "Tổng SL yêu cầu"])
        empty_suggestion = pd.DataFrame(columns=[
            "Request Number",
            "Material Number",
            "Material Description",
            "Plant",
            "Source WBS",
            "Sending Sloc",
            "Functional Location",
            "Transfer Quantity",
            "Gợi ý chuyển WBS",
        ])
        return empty_fl, empty_material, empty_plant, empty_suggestion

    summary_fl = (
        missing_stock_df.groupby("Functional Location")
        .size()
        .reset_index(name="Số dòng thiếu kho")
        .sort_values("Số dòng thiếu kho", ascending=False)
    )

    summary_material = (
        missing_stock_df.groupby(["Material Number", "Material Description"])
        .agg(
            **{
                "Số dòng thiếu kho": ("Material Number", "size"),
                "Tổng SL yêu cầu": ("Transfer Quantity", "sum"),
            }
        )
        .reset_index()
        .sort_values("Số dòng thiếu kho", ascending=False)
    )

    summary_plant = (
        missing_stock_df.groupby("Plant")
        .agg(
            **{
                "Số dòng thiếu kho": ("Plant", "size"),
                "Tổng SL yêu cầu": ("Transfer Quantity", "sum"),
            }
        )
        .reset_index()
        .sort_values("Số dòng thiếu kho", ascending=False)
    )

    suggestion = missing_stock_df[
        [
            "Request Number",
            "Material Number",
            "Material Description",
            "Plant",
            "Source WBS",
            "Sending Sloc",
            "Functional Location",
            "Transfer Quantity",
            "Gợi ý chuyển WBS",
        ]
    ].copy()

    return summary_fl, summary_material, summary_plant, suggestion


def sorted_unique_values(df: pd.DataFrame, column: str) -> list[str]:
    if column not in df.columns:
        return []
    values = df[column].dropna().astype(str)
    values = values[values.str.strip() != ""]
    return sorted(values.unique().tolist())


def apply_result_filters(df: pd.DataFrame) -> pd.DataFrame:
    filtered = df.copy()

    st.markdown('<div class="step-title">Bộ lọc báo cáo chưa đảm bảo</div>', unsafe_allow_html=True)
    with st.container():
        f1, f2, f3 = st.columns([2, 1, 1])
        keyword = f1.text_input(
            "Tìm nhanh",
            placeholder="Request, mã vật tư, mô tả, FL, WBS...",
        )
        status_filter = f2.multiselect("Tình trạng", sorted_unique_values(filtered, "Tình trạng"))
        plant_filter = f3.multiselect("Plant", sorted_unique_values(filtered, "Plant"))

        f4, f5, f6 = st.columns(3)
        fl_filter = f4.multiselect("Functional Location", sorted_unique_values(filtered, "Functional Location"))
        layer_filter = f5.multiselect("Tầng đáp ứng", sorted_unique_values(filtered, "Tầng đáp ứng"))
        sloc_filter = f6.multiselect("Sending Sloc", sorted_unique_values(filtered, "Sending Sloc"))

    if keyword:
        keyword_norm = keyword.strip().lower()
        search_cols = [
            "Request Number",
            "Material Number",
            "Material Description",
            "Functional Location",
            "Source WBS",
            "Sending Sloc",
        ]
        mask = pd.Series(False, index=filtered.index)
        for col in search_cols:
            if col in filtered.columns:
                mask = mask | filtered[col].astype(str).str.lower().str.contains(keyword_norm, na=False)
        filtered = filtered[mask]

    if status_filter:
        filtered = filtered[filtered["Tình trạng"].astype(str).isin(status_filter)]
    if plant_filter:
        filtered = filtered[filtered["Plant"].astype(str).isin(plant_filter)]
    if fl_filter:
        filtered = filtered[filtered["Functional Location"].astype(str).isin(fl_filter)]
    if layer_filter:
        filtered = filtered[filtered["Tầng đáp ứng"].astype(str).isin(layer_filter)]
    if sloc_filter:
        filtered = filtered[filtered["Sending Sloc"].astype(str).isin(sloc_filter)]

    return filtered


def auto_width_worksheet(ws) -> None:
    for col_idx, column_cells in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column_cells:
            cell_length = len(str(cell.value)) if cell.value is not None else 0
            max_length = max(max_length, cell_length)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_length + 2, 12), 48)


def format_workbook(writer, sheet_names: list[str]) -> None:
    wb = writer.book
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    bad_fill = PatternFill("solid", fgColor="FFEDD5")

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        if sheet_name == "ChiTietChuaDamBao":
            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = bad_fill
        auto_width_worksheet(ws)


def export_excel(full_df: pd.DataFrame, issue_df: pd.DataFrame, mb52_meta: Dict[str, str]) -> bytes:
    total = len(full_df)
    ok = int(full_df["Đảm bảo 100%"].sum())
    not_ok = total - ok
    error_df = full_df.loc[~full_df["Đảm bảo 100%"], DETAIL_COLUMNS].copy()
    stock_detail_df = full_df.loc[~full_df["Đảm bảo 100%"], STOCK_DETAIL_COLUMNS].copy()
    summary_fl, summary_material, summary_plant, stock_suggestion = build_stock_summaries(full_df)
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        sheet_names = ["KetLuan"]
        build_conclusion_sheet(total, ok, not_ok, mb52_meta).to_excel(writer, index=False, sheet_name="KetLuan")

        if not_ok > 0:
            error_df.to_excel(writer, index=False, sheet_name="ChiTietChuaDamBao")
            error_df[
                [
                    "Request Number",
                    "Material Number",
                    "Material Description",
                    "Plant",
                    "Functional Location",
                    "Còn thiếu",
                    "Tình trạng",
                    "Gợi ý xử lý",
                ]
            ].to_excel(writer, index=False, sheet_name="GoiYXuLy")
            stock_detail_df.to_excel(writer, index=False, sheet_name="PhanTangKho")
            summary_fl.to_excel(writer, index=False, sheet_name="TongHopThieuKho_FL")
            summary_material.to_excel(writer, index=False, sheet_name="TongHopThieuKho_VatTu")
            summary_plant.to_excel(writer, index=False, sheet_name="TongHopThieuKho_Plant")
            stock_suggestion.to_excel(writer, index=False, sheet_name="GoiYChuyenKho")
            sheet_names.extend([
                "ChiTietChuaDamBao",
                "GoiYXuLy",
                "PhanTangKho",
                "TongHopThieuKho_FL",
                "TongHopThieuKho_VatTu",
                "TongHopThieuKho_Plant",
                "GoiYChuyenKho",
            ])

        format_workbook(writer, sheet_names)

    return output.getvalue()


def render_result_card(is_all_ok: bool) -> None:
    if is_all_ok:
        st.markdown(
            """
            <div class="result-card result-ok">
                <div class="result-headline">✅ ĐẢM BẢO XUẤT KHO 100%</div>
                <div class="result-copy">Loại phiếu này đã xuất đủ toàn bộ vật tư. Không cần xử lý thêm.</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            """
            <div class="result-card result-bad">
                <div class="result-headline">⚠️ CHƯA ĐẢM BẢO XUẤT KHO 100%</div>
                <div class="result-copy">Chỉ các dòng lỗi hoặc chưa đủ được hiển thị bên dưới.</div>
            </div>
            """,
            unsafe_allow_html=True,
        )


# =====================================================
# UI
# =====================================================
st.markdown(
    f"""
    <div class="app-header">
        <div class="app-title">📦 {APP_NAME}</div>
        <div class="app-subtitle">{APP_SUBTITLE} · Version {APP_VERSION}</div>
    </div>
    """,
    unsafe_allow_html=True,
)

st.markdown('<div class="step-title">Bước 1: Chọn nguồn MB52</div>', unsafe_allow_html=True)
source_options = [
    "GitHub - MB52 mới nhất",
    "Local - data/MB52.XLSX",
    "Upload MB52 tạm thời",
]
mb52_source = st.radio("Nguồn dữ liệu MB52", source_options, horizontal=True, label_visibility="collapsed")

mb52_bytes: Optional[bytes] = None
mb52_meta: Dict[str, str] = {}

col_source, col_refresh = st.columns([4, 1])
with col_source:
    if mb52_source == "GitHub - MB52 mới nhất":
        raw_url = st.text_input("GitHub Raw URL MB52", value=get_mb52_raw_url())
        try:
            mb52_bytes, mb52_meta = download_mb52_from_github(raw_url)
        except Exception as exc:
            st.error(f"❌ Không tải được MB52 từ GitHub: {exc}")
            st.stop()
    elif mb52_source == "Local - data/MB52.XLSX":
        try:
            mb52_bytes, mb52_meta = read_local_mb52(LOCAL_MB52_PATH)
        except Exception as exc:
            st.error(f"❌ Không đọc được file local {LOCAL_MB52_PATH}: {exc}")
            st.stop()
    else:
        upload_mb52 = st.file_uploader("Upload MB52 tạm thời", type=["xlsx", "xls"], key="mb52_upload")
        if not upload_mb52:
            st.info("Vui lòng upload file MB52 để tiếp tục.")
            st.stop()
        mb52_bytes = upload_mb52.getvalue()
        mb52_meta = {
            "source": "Upload MB52 tạm thời",
            "url": upload_mb52.name,
            "loaded_at": datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
            "last_modified": "",
            "etag": "",
        }

with col_refresh:
    st.write("")
    st.write("")
    if st.button("🔄 Làm mới MB52", use_container_width=True):
        st.cache_data.clear()
        st.rerun()

mb52_raw = load_mb52(mb52_bytes)
st.success(
    f"Đã sẵn sàng MB52: {len(mb52_raw):,} dòng · "
    f"{mb52_raw['Material'].nunique():,} mã vật tư · "
    f"nguồn {mb52_meta.get('source', '')}"
)

st.markdown('<div class="step-title">Bước 2: Upload phiếu xuất kho</div>', unsafe_allow_html=True)
issue_file = st.file_uploader(
    "Chọn file phiếu xuất kho",
    type=["xlsx", "xls"],
    help="File cần có Transfer Quantity, Actual Quantity ở cột AB hoặc theo tên cột, và Status ở cột AC hoặc theo tên cột.",
)
if not issue_file:
    st.info("Upload phiếu xuất kho để phần mềm kết luận ngay.")
    st.stop()

issue_df = load_issue(issue_file.getvalue())

with st.spinner("Đang kiểm tra trạng thái thực xuất và tồn kho MB52 theo 5 tầng..."):
    stock_report = build_sequential_5_layer(issue_df, mb52_raw)
    final_report = build_business_conclusion(stock_report)

total_lines = len(final_report)
ok_lines = int(final_report["Đảm bảo 100%"].sum())
not_ok_lines = total_lines - ok_lines
ok_rate = (ok_lines / total_lines * 100) if total_lines else 0
is_all_ok = total_lines > 0 and not_ok_lines == 0

st.markdown('<div class="step-title">Bước 3: Xem kết luận</div>', unsafe_allow_html=True)
metric1, metric2, metric3, metric4 = st.columns(4)
metric1.metric("Tổng dòng", f"{total_lines:,}")
metric2.metric("Đã xuất đủ", f"{ok_lines:,}")
metric3.metric("Chưa đảm bảo", f"{not_ok_lines:,}")
metric4.metric("Tỷ lệ đảm bảo", f"{ok_rate:.1f}%")

render_result_card(is_all_ok)

if not is_all_ok:
    not_ok_report = final_report.loc[~final_report["Đảm bảo 100%"]].copy()
    filtered_not_ok_report = apply_result_filters(not_ok_report)

    st.caption(f"Đang hiển thị {len(filtered_not_ok_report):,}/{len(not_ok_report):,} dòng chưa đảm bảo theo bộ lọc hiện tại.")

    if filtered_not_ok_report.empty:
        st.info("Không có dòng nào khớp bộ lọc hiện tại.")
    else:
        error_df = filtered_not_ok_report[DETAIL_COLUMNS].copy()
        stock_detail_df = filtered_not_ok_report[STOCK_DETAIL_COLUMNS].copy()

        error_counts = error_df["Tình trạng"].value_counts().rename_axis("Tình trạng").reset_index(name="Số dòng")
        st.dataframe(error_counts, use_container_width=True, hide_index=True, height=150)

        st.markdown('<div class="step-title">Các dòng cần xử lý</div>', unsafe_allow_html=True)
        st.dataframe(
            error_df,
            use_container_width=True,
            hide_index=True,
            height=430,
            column_config={
                "Transfer Quantity": st.column_config.NumberColumn("Transfer Quantity", format="%.2f"),
                "Actual Quantity": st.column_config.NumberColumn("Actual Quantity", format="%.2f"),
                "Còn thiếu": st.column_config.NumberColumn("Còn thiếu", format="%.2f"),
                "Gợi ý xử lý": st.column_config.TextColumn("Gợi ý xử lý", width="large"),
            },
        )

        st.markdown('<div class="step-title">Báo cáo tính toán chuyển kho / chuyển dự án</div>', unsafe_allow_html=True)
        with st.expander("Phân tầng kho và gợi ý chuyển kho", expanded=True):
            st.dataframe(
                stock_detail_df,
                use_container_width=True,
                hide_index=True,
                height=430,
                column_config={
                    "Transfer Quantity": st.column_config.NumberColumn("Transfer Quantity", format="%.2f"),
                    "Actual Quantity": st.column_config.NumberColumn("Actual Quantity", format="%.2f"),
                    "Còn thiếu": st.column_config.NumberColumn("Còn thiếu", format="%.2f"),
                    "Tồn kho DA CN": st.column_config.NumberColumn("Tồn kho DA CN", format="%.2f"),
                    "Tồn kho DA Tỉnh": st.column_config.NumberColumn("Tồn kho DA Tỉnh", format="%.2f"),
                    "Tồn kho CN": st.column_config.NumberColumn("Tồn kho CN", format="%.2f"),
                    "Tồn kho Tỉnh": st.column_config.NumberColumn("Tồn kho Tỉnh", format="%.2f"),
                    "Tồn kho Khu vực": st.column_config.NumberColumn("Tồn kho Khu vực", format="%.2f"),
                    "Gợi ý chuyển WBS": st.column_config.TextColumn("Gợi ý chuyển WBS", width="large"),
                },
            )

            summary_fl, summary_material, summary_plant, stock_suggestion = build_stock_summaries(filtered_not_ok_report)
            tab_fl, tab_material, tab_plant, tab_suggestion = st.tabs([
                "Theo FL",
                "Theo vật tư",
                "Theo Plant",
                "Gợi ý chuyển kho",
            ])
            with tab_fl:
                st.dataframe(summary_fl, use_container_width=True, hide_index=True, height=260)
            with tab_material:
                st.dataframe(summary_material, use_container_width=True, hide_index=True, height=260)
            with tab_plant:
                st.dataframe(summary_plant, use_container_width=True, hide_index=True, height=260)
            with tab_suggestion:
                st.dataframe(stock_suggestion, use_container_width=True, hide_index=True, height=260)

export_bytes = export_excel(final_report, issue_df, mb52_meta)
file_time = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
st.download_button(
    label="⬇️ Tải kết quả Excel",
    data=export_bytes,
    file_name=f"StockFlow_KetQua_XuatKho_{file_time}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
)

st.caption("StockFlow Checker · Người dùng upload phiếu, phần mềm trả lời ngay: đảm bảo 100% hoặc thiếu dòng nào, vì sao, xử lý thế nào.")
